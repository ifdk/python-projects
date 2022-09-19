import openpyxl
from pathlib import Path

xlsx_file = Path(Path.home(), 'Documents', 'Python', 'ashburntemp.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)

# Read the active sheet:
sheet = wb_obj.active
for row in sheet.iter_rows(max_row=13):
    for cell in row:
        print(cell.value, end=" ")
    print()

print(sheet["C2"].value)
print(sheet.max_row, sheet.max_column)